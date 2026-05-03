"""
Val Prop Tracker — Excel workbook generator and result updater.
Mirrors the MMA analyzer tracker pattern:
  • val auto      → appends new rows each run
  • val tracker   → rebuilds entire workbook from all saved JSON reports
  • val results   → writes actual outcomes non-destructively

Workbook has 3 sheets:
  1. Props       — one row per player/stat simulation with edge data
  2. Best Plays  — top-EV plays filtered from Props (formula-driven)
  3. Performance — per-player hit-rate and ROI roll-up (formula-driven)
"""

from __future__ import annotations

import json
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
_NAVY       = "1B2A4A"
_GOLD       = "C9A84C"
_RED_VAL    = "C8102E"        # Valorant brand red
_WHITE      = "FFFFFF"
_LIGHT_GREY = "F5F5F5"
_MID_GREY   = "E0E0E0"
_GREEN_BG   = "D6F0D6"
_GREEN_FG   = "1A6B1A"
_RED_BG     = "FAD7D7"
_RED_FG     = "8B1A1A"
_AMBER_BG   = "FFF3CD"
_AMBER_FG   = "7A5C00"
_INPUT_FILL = "FFFBE6"        # yellow tint for user-input cells

_TIER_FILLS = {
    "ELITE":    PatternFill("solid", fgColor="D6F0D6"),
    "STRONG":   PatternFill("solid", fgColor="EBF5D6"),
    "PLAYABLE": PatternFill("solid", fgColor="FFF3CD"),
    "PASS":     PatternFill("solid", fgColor="FAD7D7"),
}

# Player stripe colours (cycle)
_PLAYER_FILLS = [
    PatternFill("solid", fgColor="EEF3FB"),  # soft blue
    PatternFill("solid", fgColor="F3FAEE"),  # soft green
    PatternFill("solid", fgColor="FBF3EE"),  # soft orange
    PatternFill("solid", fgColor="F8EEFB"),  # soft purple
    PatternFill("solid", fgColor="EEFBF8"),  # soft teal
    PatternFill("solid", fgColor="FBEEF3"),  # soft pink
    PatternFill("solid", fgColor="F3EEFB"),  # soft lavender
    PatternFill("solid", fgColor="FAFBEE"),  # soft yellow
    PatternFill("solid", fgColor="EEF8FB"),  # soft cyan
    PatternFill("solid", fgColor="FBEEE8"),  # soft salmon
]

# ── Style helpers ─────────────────────────────────────────────────────────────

def _hfont()  -> Font: return Font(bold=True, color=_WHITE, name="Calibri", size=10)
def _tfont()  -> Font: return Font(bold=True, color=_NAVY,  name="Calibri", size=13)
def _bfont()  -> Font: return Font(bold=True,               name="Calibri", size=10)
def _nfont(c: str = "000000") -> Font: return Font(name="Calibri", size=10, color=c)

def _hfill()  -> PatternFill: return PatternFill("solid", fgColor=_NAVY)
def _gfill()  -> PatternFill: return PatternFill("solid", fgColor=_GOLD)
def _rfill()  -> PatternFill: return PatternFill("solid", fgColor=_RED_VAL)

def _border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _center() -> Alignment: return Alignment(horizontal="center", vertical="center")
def _left()   -> Alignment: return Alignment(horizontal="left",   vertical="center")

def _col(ws, idx: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(idx)].width = width

def _cell(ws, row: int, col: int, value: Any, *,
          font=None, fill=None, align=None, fmt: str = None) -> None:
    c = ws.cell(row=row, column=col, value=value)
    if font:  c.font      = font
    if fill:  c.fill      = fill
    if align: c.alignment = align
    c.border = _border()
    if fmt:   c.number_format = fmt

def _header_row(ws, row: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, 1):
        _cell(ws, row, i, h, font=_hfont(), fill=_hfill(), align=_center())


# ── Column spec for Props sheet ───────────────────────────────────────────────
#   (header, width, number_format or None)
_PROPS_COLS = [
    ("Run Date",      14, None),          # 1
    ("Player",        13, None),          # 2
    ("Stat",          10, None),          # 3
    ("Line",           8, "0.0"),         # 4
    ("Best Side",     10, None),          # 5
    ("Expected",      10, "0.00"),        # 6
    ("Win %",          8, "0.0%"),        # 7
    ("Implied %",     10, "0.0%"),        # 8
    ("Edge",           8, '"+0.0%;-0.0%;"0.0%"'), # 9
    ("EV",             8, "+0.000;-0.000"),  # 10
    ("Kelly",          7, "0.000"),        # 11
    ("Tier",          10, None),           # 12
    ("Confidence",    11, None),           # 13
    ("Volatility",    10, None),           # 14
    ("5th Pct",        8, "0.00"),         # 15
    ("95th Pct",       8, "0.00"),         # 16
    ("Sims",           8, "#,##0"),        # 17
    ("Actual Result", 15, None),           # 18  ← user fills: OVER / UNDER
    ("Hit?",           8, None),           # 19  ← formula
    ("Notes",         22, None),           # 20
]

_COL_ACTUAL   = 18
_COL_HIT      = 19
_COL_SIDE     = 5
_COL_TIER     = 12


# ── Data model ────────────────────────────────────────────────────────────────

@dataclass
class PropRow:
    run_date:    str
    player:      str
    stat:        str
    line:        float
    best_side:   str       # "over" | "under"
    ev:          float
    p_win:       float     # 0–100
    implied_pct: float     # 0–100
    edge:        float     # percentage points
    ev_bet:      float
    kelly:       float
    tier:        str
    confidence:  str
    volatility:  str
    p5:          float
    p95:         float
    n_sims:      int
    actual:      str = ""  # user-filled: OVER / UNDER
    notes:       str = ""

    @property
    def unique_key(self) -> tuple:
        """Stable identity for non-destructive result updates."""
        return (self.run_date, self.player, self.stat, self.best_side)


def results_to_rows(results: list[dict], run_date: str) -> list[PropRow]:
    rows = []
    for res in results:
        player = res["player"]
        for stat, d in res["stats"].items():
            side  = d["best_side"]
            p_win = d["p_over"] if side == "over" else d["p_under"]
            rows.append(PropRow(
                run_date    = run_date,
                player      = player,
                stat        = stat,
                line        = d["line"],
                best_side   = side.upper(),
                ev          = d["ev"],
                p_win       = p_win,
                implied_pct = d["implied_prob"],
                edge        = d["best_edge"],
                ev_bet      = d["best_ev"],
                kelly       = d["best_kelly"],
                tier        = d["best_tier"],
                confidence  = d["confidence"],
                volatility  = d["volatility"],
                p5          = d["p5"],
                p95         = d["p95"],
                n_sims      = res["n_sims"],
            ))
    return rows


# ── Sheet 1: Props ─────────────────────────────────────────────────────────────

def _build_props_sheet(ws, rows: list[PropRow]) -> None:
    ws.title = "Props"
    ws.freeze_panes = "A2"

    # Header
    _header_row(ws, 1, [c[0] for c in _PROPS_COLS])
    for i, (_, w, _) in enumerate(_PROPS_COLS, 1):
        _col(ws, i, w)

    # Player → fill colour map
    player_fills: dict[str, PatternFill] = {}
    fill_idx = 0

    for dr, row in enumerate(rows, start=2):
        if row.player not in player_fills:
            player_fills[row.player] = _PLAYER_FILLS[fill_idx % len(_PLAYER_FILLS)]
            fill_idx += 1
        pf = player_fills[row.player]

        def c(col: int, val: Any, fmt: str = None, align=None, font=None, fill=None):
            _cell(ws, dr, col, val,
                  font  = font or _nfont(),
                  fill  = fill or pf,
                  align = align or _left(),
                  fmt   = fmt)

        c(1,  row.run_date)
        c(2,  row.player,     font=_bfont())
        c(3,  row.stat.upper() if row.stat in ("acs","adr","kd") else row.stat.title())
        c(4,  row.line,       fmt="0.0",         align=_center())
        # Side coloured
        side_font = Font(bold=True, color=_GREEN_FG if row.best_side == "OVER" else _RED_FG,
                         name="Calibri", size=10)
        c(5,  row.best_side,  font=side_font,    align=_center())
        c(6,  row.ev,         fmt="0.00",         align=_center())
        c(7,  row.p_win / 100, fmt="0.0%",        align=_center())
        c(8,  row.implied_pct / 100, fmt="0.0%",  align=_center())

        # Edge — colour by tier
        edge_fill  = _TIER_FILLS.get(row.tier, pf)
        edge_color = (_GREEN_FG if row.edge > 5 else
                      _AMBER_FG if row.edge > 3 else
                      _RED_FG)
        edge_str = f"+{row.edge:.1f}%" if row.edge > 0 else f"{row.edge:.1f}%"
        c(9,  edge_str, fill=edge_fill, align=_center(),
          font=Font(bold=True, color=edge_color, name="Calibri", size=10))

        ev_str = f"+{row.ev_bet:.3f}" if row.ev_bet > 0 else f"{row.ev_bet:.3f}"
        c(10, ev_str,   align=_center())
        c(11, row.kelly, fmt="0.000",              align=_center())

        tier_fill = _TIER_FILLS.get(row.tier, pf)
        tier_font = Font(bold=True, name="Calibri", size=10,
                         color=(_GREEN_FG if row.tier in ("ELITE","STRONG")
                                else _AMBER_FG if row.tier == "PLAYABLE"
                                else _RED_FG))
        c(12, row.tier,  fill=tier_fill, align=_center(), font=tier_font)
        c(13, row.confidence, align=_center())
        c(14, row.volatility, align=_center())
        c(15, row.p5,    fmt="0.00", align=_center())
        c(16, row.p95,   fmt="0.00", align=_center())
        c(17, row.n_sims, fmt="#,##0", align=_center())

        # Actual Result — highlighted input cell
        _cell(ws, dr, _COL_ACTUAL, row.actual or "",
              font=_nfont(), fill=PatternFill("solid", fgColor=_INPUT_FILL),
              align=_center())

        # Hit? — formula =IF(R2="","—",IF(R2=E2,"HIT","MISS"))
        col_actual_ltr = get_column_letter(_COL_ACTUAL)
        col_side_ltr   = get_column_letter(_COL_SIDE)
        formula = (f'=IF({col_actual_ltr}{dr}="",'
                   f'"—",'
                   f'IF({col_actual_ltr}{dr}={col_side_ltr}{dr},"HIT","MISS"))')
        hit_cell = ws.cell(row=dr, column=_COL_HIT, value=formula)
        hit_cell.font      = _bfont()
        hit_cell.alignment = _center()
        hit_cell.border    = _border()
        # Conditional fill handled below via data_bar / manual check — skip for now

        c(20, row.notes)

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_PROPS_COLS))}1"


# ── Sheet 2: Best Plays ───────────────────────────────────────────────────────

def _build_best_plays_sheet(ws, rows: list[PropRow]) -> None:
    ws.title = "Best Plays"
    ws.freeze_panes = "A3"

    # Title bar
    ws.merge_cells("A1:K1")
    title = ws.cell(row=1, column=1, value="VAL PROP ENGINE — BEST PLAYS")
    title.font      = Font(bold=True, color=_WHITE, name="Calibri", size=14)
    title.fill      = PatternFill("solid", fgColor=_RED_VAL)
    title.alignment = _center()
    title.border    = _border()
    for col in range(2, 12):
        ws.cell(row=1, column=col).fill   = PatternFill("solid", fgColor=_RED_VAL)
        ws.cell(row=1, column=col).border = _border()

    bp_cols = [
        ("Rank", 6), ("Player", 13), ("Stat", 10), ("Side", 9),
        ("Line", 8), ("Edge", 9), ("EV", 9), ("Kelly", 8),
        ("Win %", 8), ("Tier", 10), ("Confidence", 12),
    ]
    for i, (h, w) in enumerate(bp_cols, 1):
        _cell(ws, 2, i, h, font=_hfont(), fill=_hfill(), align=_center())
        _col(ws, i, w)

    # Filter: exclude PASS, sort by ev_bet desc
    eligible = sorted(
        [r for r in rows if r.tier != "PASS"],
        key=lambda r: r.ev_bet, reverse=True
    )

    for rank, row in enumerate(eligible[:20], 1):
        pf        = _TIER_FILLS.get(row.tier, PatternFill("solid", fgColor=_LIGHT_GREY))
        edge_str  = f"+{row.edge:.1f}%" if row.edge > 0 else f"{row.edge:.1f}%"
        ev_str    = f"+{row.ev_bet:.3f}" if row.ev_bet > 0 else f"{row.ev_bet:.3f}"
        side_col  = _GREEN_FG if row.best_side == "OVER" else _RED_FG

        def bp(col: int, val: Any, align=_center(), font=None, fill=None):
            _cell(ws, rank + 2, col, val,
                  font=font or _nfont(), fill=fill or pf, align=align)

        bp(1,  rank,     font=_bfont())
        bp(2,  row.player, font=_bfont(), align=_left())
        bp(3,  row.stat.upper() if row.stat in ("acs","adr","kd") else row.stat.title(),
           align=_left())
        bp(4,  row.best_side,
           font=Font(bold=True, color=side_col, name="Calibri", size=10))
        bp(5,  row.line)
        bp(6,  edge_str,
           font=Font(bold=True, color=(_GREEN_FG if row.edge > 5 else _AMBER_FG),
                     name="Calibri", size=10),
           fill=pf)
        bp(7,  ev_str,
           font=Font(bold=True, color=(_GREEN_FG if row.ev_bet > 0.05 else _AMBER_FG),
                     name="Calibri", size=10))
        bp(8,  f"{row.kelly:.3f}")
        bp(9,  f"{row.p_win:.1f}%")
        bp(10, row.tier,
           font=Font(bold=True, name="Calibri", size=10,
                     color=(_GREEN_FG if row.tier in ("ELITE","STRONG")
                            else _AMBER_FG)),
           fill=pf)
        bp(11, row.confidence)


# ── Sheet 3: Performance ──────────────────────────────────────────────────────

def _build_performance_sheet(ws, rows: list[PropRow]) -> None:
    ws.title = "Performance"
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:I1")
    title = ws.cell(row=1, column=1, value="VAL PROP ENGINE — PERFORMANCE TRACKER")
    title.font      = Font(bold=True, color=_WHITE, name="Calibri", size=14)
    title.fill      = PatternFill("solid", fgColor=_NAVY)
    title.alignment = _center()
    title.border    = _border()
    for col in range(2, 10):
        ws.cell(row=1, column=col).fill   = PatternFill("solid", fgColor=_NAVY)
        ws.cell(row=1, column=col).border = _border()

    perf_cols = [
        ("Player", 14), ("Total Props", 12), ("Resolved", 10),
        ("Hits", 8), ("Misses", 8), ("Hit Rate", 10),
        ("Avg Edge", 10), ("Avg EV", 10), ("Avg Kelly", 10),
    ]
    for i, (h, w) in enumerate(perf_cols, 1):
        _cell(ws, 2, i, h, font=_hfont(), fill=_hfill(), align=_center())
        _col(ws, i, w)

    # Roll up by player
    players = list(dict.fromkeys(r.player for r in rows))  # preserve order
    for dr, player in enumerate(players, start=3):
        pr     = [r for r in rows if r.player == player]
        total  = len(pr)
        res    = [r for r in pr if r.actual.strip().upper() in ("OVER", "UNDER")]
        hits   = sum(1 for r in res if r.actual.strip().upper() == r.best_side)
        misses = len(res) - hits
        hit_rt = hits / len(res) if res else 0.0
        avg_edge = sum(r.edge for r in pr) / total if total else 0
        avg_ev   = sum(r.ev_bet for r in pr) / total if total else 0
        avg_kelly= sum(r.kelly  for r in pr) / total if total else 0

        pf = _PLAYER_FILLS[players.index(player) % len(_PLAYER_FILLS)]

        def pc(col: int, val: Any, fmt: str = None):
            _cell(ws, dr, col, val, font=_nfont(), fill=pf,
                  align=_center(), fmt=fmt)

        pc(1,  player);                        ws.cell(dr, 1).font      = _bfont()
        ws.cell(dr, 1).alignment = _left()
        pc(2,  total)
        pc(3,  len(res))
        pc(4,  hits,   fmt="0")
        pc(5,  misses, fmt="0")
        pc(6,  hit_rt, fmt="0.0%")
        edge_str = f"+{avg_edge:.1f}%" if avg_edge > 0 else f"{avg_edge:.1f}%"
        ev_str   = f"+{avg_ev:.3f}"   if avg_ev   > 0 else f"{avg_ev:.3f}"
        pc(7,  edge_str)
        pc(8,  ev_str)
        pc(9,  f"{avg_kelly:.3f}")

    # OVERALL row
    dr_total = len(players) + 3
    all_res  = [r for r in rows if r.actual.strip().upper() in ("OVER", "UNDER")]
    all_hits = sum(1 for r in all_res if r.actual.strip().upper() == r.best_side)
    overall_rt = all_hits / len(all_res) if all_res else 0.0
    overall_edge = sum(r.edge   for r in rows) / len(rows) if rows else 0
    overall_ev   = sum(r.ev_bet for r in rows) / len(rows) if rows else 0
    overall_kelly= sum(r.kelly  for r in rows) / len(rows) if rows else 0

    nav_fill = PatternFill("solid", fgColor=_NAVY)
    vals = [
        "OVERALL", len(rows), len(all_res), all_hits, len(all_res) - all_hits,
        f"{overall_rt:.1%}",
        f"+{overall_edge:.1f}%" if overall_edge > 0 else f"{overall_edge:.1f}%",
        f"+{overall_ev:.3f}"   if overall_ev   > 0 else f"{overall_ev:.3f}",
        f"{overall_kelly:.3f}",
    ]
    for i, v in enumerate(vals, 1):
        _cell(ws, dr_total, i, v,
              font=Font(bold=True, color=_WHITE, name="Calibri", size=10),
              fill=nav_fill, align=_center())

    # Instructions block
    inst_row = dr_total + 2
    ws.cell(row=inst_row, column=1,
            value="HOW TO UPDATE RESULTS").font = Font(bold=True, name="Calibri", size=11)
    instructions = [
        "1. Go to the Props sheet",
        "2. In the 'Actual Result' column (col R), type OVER or UNDER for each resolved prop",
        "3. The 'Hit?' column auto-calculates: HIT if correct, MISS if wrong",
        "4. Or run:  val results --player TenZ --stat kills --side over --result hit",
        "5. For bulk updates:  val results --file my_results.json",
        "6. Refresh this sheet by re-running:  val tracker",
    ]
    for i, line in enumerate(instructions, start=inst_row + 1):
        ws.cell(row=i, column=1, value=line).font = Font(name="Calibri", size=10, color="444444")


# ── Workbook builder ──────────────────────────────────────────────────────────

def build_prop_workbook(rows: list[PropRow]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_props = wb.create_sheet("Props")
    ws_best  = wb.create_sheet("Best Plays")
    ws_perf  = wb.create_sheet("Performance")

    _build_props_sheet(ws_props, rows)
    _build_best_plays_sheet(ws_best, rows)
    _build_performance_sheet(ws_perf, rows)

    return wb


# ── Non-destructive append ────────────────────────────────────────────────────

def _read_existing_rows(xlsx_path: Path) -> list[PropRow]:
    """Load all existing rows from the Props sheet, preserving actual/notes."""
    if not xlsx_path.exists():
        return []
    try:
        wb   = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws   = wb["Props"]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r[0]:
                continue
            try:
                rows.append(PropRow(
                    run_date    = str(r[0]),
                    player      = str(r[1]),
                    stat        = str(r[2]).lower(),
                    line        = float(r[3]),
                    best_side   = str(r[4]),
                    ev          = float(r[5]) if r[5] is not None else 0.0,
                    p_win       = float(str(r[6]).replace("%","")) if r[6] else 0.0,
                    implied_pct = float(str(r[7]).replace("%","")) if r[7] else 52.38,
                    edge        = float(str(r[8]).replace("+","").replace("%","").strip()) if r[8] else 0.0,
                    ev_bet      = float(str(r[9]).replace("+","").strip()) if r[9] else 0.0,
                    kelly       = float(r[10]) if r[10] is not None else 0.0,
                    tier        = str(r[11]),
                    confidence  = str(r[12]),
                    volatility  = str(r[13]),
                    p5          = float(r[14]) if r[14] is not None else 0.0,
                    p95         = float(r[15]) if r[15] is not None else 0.0,
                    n_sims      = int(r[16]) if r[16] is not None else 0,
                    actual      = str(r[17]) if r[17] else "",
                    notes       = str(r[19]) if r[19] else "",
                ))
            except Exception:
                continue
        return rows
    except Exception:
        return []


def append_to_tracker(xlsx_path: Path, new_rows: list[PropRow]) -> int:
    """
    Merges new_rows into the existing tracker.
    Existing rows (matched by unique_key) are preserved;
    their actual/notes fields are kept.
    New rows are appended.
    Returns number of new rows added.
    """
    existing  = _read_existing_rows(xlsx_path)
    exist_map = {r.unique_key: r for r in existing}

    added = 0
    for nr in new_rows:
        if nr.unique_key not in exist_map:
            exist_map[nr.unique_key] = nr
            added += 1
        else:
            # Keep actual/notes from existing record
            ex = exist_map[nr.unique_key]
            nr.actual = ex.actual
            nr.notes  = ex.notes
            exist_map[nr.unique_key] = nr

    all_rows = list(exist_map.values())
    all_rows.sort(key=lambda r: (r.run_date, r.player, r.stat))

    wb = build_prop_workbook(all_rows)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return added


# ── Result updater ────────────────────────────────────────────────────────────

def update_results(xlsx_path: Path,
                   entries: list[dict]) -> tuple[int, list[str]]:
    """
    Non-destructively writes 'Actual Result' into the Props sheet.
    Each entry: {player, stat, side, result}  where result = "over" | "under"
    Matches rows by (player, stat, side); updates first unresolved match.
    Returns (n_updated, unmatched_labels).
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Tracker not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Props"]
    updated   = 0
    unmatched = []

    for entry in entries:
        player = str(entry.get("player", "")).strip().lower()
        stat   = str(entry.get("stat",   "")).strip().lower()
        side   = str(entry.get("side",   "")).strip().upper()
        result = str(entry.get("result", "")).strip().upper()

        if result not in ("OVER", "UNDER"):
            unmatched.append(f"{player}/{stat}/{side} — bad result '{result}'")
            continue

        found = False
        for row in ws.iter_rows(min_row=2):
            r_player = str(row[1].value or "").strip().lower()
            r_stat   = str(row[2].value or "").strip().lower()
            r_side   = str(row[4].value or "").strip().upper()
            r_actual = str(row[_COL_ACTUAL - 1].value or "").strip()

            if (r_player == player and r_stat == stat and
                    r_side == side and not r_actual):
                row[_COL_ACTUAL - 1].value = result
                updated += 1
                found = True
                break

        if not found:
            unmatched.append(f"{entry.get('player')}/{stat}/{side}")

    wb.save(xlsx_path)
    return updated, unmatched


# ── Tracker rebuild from JSON dir ─────────────────────────────────────────────

def generate_tracker(json_dir: Path, xlsx_path: Path) -> tuple[int, int]:
    """
    Scans json_dir for report_*.json files, parses them all,
    and (re)builds the tracker workbook, preserving any existing actual/notes.
    Returns (n_reports, n_rows).
    """
    json_files = sorted(json_dir.glob("report_*.json"))
    if not json_files:
        return 0, 0

    all_new_rows: list[PropRow] = []
    for jf in json_files:
        try:
            with open(jf, encoding="utf-8") as f:
                data = json.load(f)
            run_date = jf.stem.replace("report_", "")
            # Normalise run_date: 20260503_155022 → 2026-05-03 15:50
            if len(run_date) >= 15:
                d = run_date
                run_date = f"{d[0:4]}-{d[4:6]}-{d[6:8]} {d[9:11]}:{d[11:13]}"
            all_new_rows.extend(results_to_rows(data, run_date))
        except Exception:
            continue

    if not all_new_rows:
        return len(json_files), 0

    # Preserve existing actual/notes
    existing  = _read_existing_rows(xlsx_path)
    exist_map = {r.unique_key: r for r in existing}

    for nr in all_new_rows:
        if nr.unique_key in exist_map:
            ex = exist_map[nr.unique_key]
            nr.actual = ex.actual
            nr.notes  = ex.notes

    all_new_rows.sort(key=lambda r: (r.run_date, r.player, r.stat))
    wb = build_prop_workbook(all_new_rows)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return len(json_files), len(all_new_rows)


# ── List unresolved ───────────────────────────────────────────────────────────

def list_unresolved(xlsx_path: Path) -> list[dict]:
    rows = _read_existing_rows(xlsx_path)
    return [
        {"player": r.player, "stat": r.stat, "side": r.best_side,
         "line": r.line, "tier": r.tier, "ev": r.ev_bet, "run_date": r.run_date}
        for r in rows
        if not r.actual.strip()
    ]
